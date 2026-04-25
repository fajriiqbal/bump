from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.models import PondokProfile
from finance.models import Bill


def _student_rows():
    bills = Bill.objects.select_related("santri", "jenis_pembayaran").order_by(
        "santri__nama_lengkap",
        "periode_tahun",
        "periode_bulan",
        "id",
    )
    grouped = {}
    for bill in bills:
        bucket = grouped.setdefault(
            bill.santri_id,
            {
                "santri": bill.santri.nama_lengkap,
                "nis": bill.santri.nis,
                "kelas": bill.santri.kelas or "-",
                "wali": bill.santri.nama_wali or "-",
                "bills": [],
            },
        )
        bucket["bills"].append(bill)

    rows = []
    for data in grouped.values():
        bills_for_student = data["bills"]
        latest_bill = bills_for_student[-1]
        latest_due = latest_bill.jatuh_tempo
        status_order = {
            Bill.Status.TERLAMBAT: 3,
            Bill.Status.SEBAGIAN: 2,
            Bill.Status.BELUM: 1,
            Bill.Status.LUNAS: 0,
        }
        effective_status = max(bills_for_student, key=lambda bill: status_order.get(bill.effective_status, 0)).effective_status
        status_label = max(bills_for_student, key=lambda bill: status_order.get(bill.effective_status, 0)).effective_status_label
        if effective_status == Bill.Status.TERLAMBAT:
            status_label = "Terlambat"
        elif effective_status == Bill.Status.SEBAGIAN:
            status_label = "Sebagian"
        elif effective_status == Bill.Status.BELUM:
            status_label = "Belum Bayar"
        else:
            status_label = "Lunas"

        detail_lines = []
        for bill in bills_for_student:
            detail_lines.append(
                f"{bill.jenis_pembayaran.nama} {bill.periode_bulan:02d}/{bill.periode_tahun} | "
                f"Tagihan {bill.total_tagihan:,} | Dibayar {bill.total_dibayar:,} | "
                f"Sisa {bill.sisa_tagihan:,} | {bill.effective_status_label}"
            )

        rows.append(
            {
                "santri": data["santri"],
                "nis": data["nis"],
                "kelas": data["kelas"],
                "wali": data["wali"],
                "bill_count": len(bills_for_student),
                "total_tagihan": sum((bill.total_tagihan for bill in bills_for_student), 0),
                "total_dibayar": sum((bill.total_dibayar for bill in bills_for_student), 0),
                "total_sisa": sum((bill.sisa_tagihan for bill in bills_for_student), 0),
                "latest_due": latest_due,
                "status": status_label,
                "detail": "\n".join(detail_lines),
            }
        )
    return rows


def _summary_rows(rows):
    total_tagihan = sum((row["total_tagihan"] for row in rows), 0)
    total_dibayar = sum((row["total_dibayar"] for row in rows), 0)
    total_sisa = sum((row["total_sisa"] for row in rows), 0)
    overdue = sum(1 for row in rows if row["status"] == "Terlambat")
    paid = sum(1 for row in rows if row["status"] == "Lunas")
    return [
        ("Total tagihan", total_tagihan),
        ("Total dibayar", total_dibayar),
        ("Sisa tagihan", total_sisa),
        ("Lunas", paid),
        ("Terlambat", overdue),
        ("Jumlah santri", len(rows)),
    ]


def build_bills_workbook():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    rows = _student_rows()
    profile = PondokProfile.get_solo()
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Ringkasan"
    ws_detail = wb.create_sheet("Tagihan")

    header_fill = PatternFill("solid", fgColor="A05A2C")
    soft_fill = PatternFill("solid", fgColor="FFF8EF")
    thin = Side(style="thin", color="D6C8B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_summary["A1"] = "Laporan Tagihan"
    ws_summary["A1"].font = Font(size=15, bold=True, color="3E2A1F")
    ws_summary["A2"] = f"Digenerate: {timezone.localtime(timezone.now()).strftime('%d %b %Y %H:%M')}"
    ws_summary["A2"].font = Font(size=10, color="6B5B4E")
    ws_summary["A3"] = profile.display_name or "Pondok Pesantren"
    ws_summary["A3"].font = Font(size=11, bold=True, color="3E2A1F")
    ws_summary["A4"] = profile.display_address or "Identitas pondok belum dilengkapi."
    ws_summary["A4"].font = Font(size=10, color="6B5B4E")
    ws_summary["A5"] = " | ".join(part for part in [profile.telepon, profile.wa_admin, profile.email] if part) or "-"
    ws_summary["A5"].font = Font(size=10, color="6B5B4E")

    ws_summary["A7"] = "Ikhtisar"
    ws_summary["A7"].font = Font(size=12, bold=True, color="3E2A1F")

    summary_rows = _summary_rows(rows)
    for idx, (label, value) in enumerate(summary_rows, start=8):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"B{idx}"] = value
        ws_summary[f"A{idx}"].fill = soft_fill
        ws_summary[f"B{idx}"].fill = soft_fill
        ws_summary[f"A{idx}"].border = border
        ws_summary[f"B{idx}"].border = border
        ws_summary[f"A{idx}"].alignment = Alignment(horizontal="left")
        ws_summary[f"B{idx}"].alignment = Alignment(horizontal="right")

    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 18

    headers = ["Santri", "NIS", "Kelas", "Wali", "Jumlah Tagihan", "Total Tagihan", "Total Dibayar", "Sisa", "Jatuh Tempo Terakhir", "Status", "Rincian Tagihan"]
    ws_detail.append(headers)
    for cell in ws_detail[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in rows:
        ws_detail.append(
            [
                row["santri"],
                row["nis"],
                row["kelas"],
                row["wali"],
                row["bill_count"],
                row["total_tagihan"],
                row["total_dibayar"],
                row["total_sisa"],
                row["latest_due"],
                row["status"],
                row["detail"],
            ]
        )

    for row in ws_detail.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
        row[4].number_format = '0'
        row[5].number_format = '#,##0'
        row[6].number_format = '#,##0'
        row[7].number_format = '#,##0'
        row[8].number_format = "dd-mm-yyyy"
        row[10].alignment = Alignment(wrap_text=True, vertical="top")

    widths = [26, 16, 14, 24, 16, 16, 16, 16, 18, 14, 52]
    for column, width in zip("ABCDEFGHIJK", widths):
        ws_detail.column_dimensions[column].width = width

    ws_detail.freeze_panes = "A2"
    ws_detail.auto_filter.ref = ws_detail.dimensions

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_bills_pdf():
    rows = _student_rows()
    profile = PondokProfile.get_solo()
    logo_path = Path(settings.BASE_DIR) / "static" / "img" / "pondok-logo.png"
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#3E2A1F"),
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "ReportMeta",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#6B5B4E"),
        spaceAfter=8,
    )

    story = [
        Image(str(logo_path), width=24 * mm, height=24 * mm) if logo_path.exists() else Spacer(1, 2 * mm),
        Spacer(1, 2 * mm),
        Paragraph(profile.display_name or "Pondok Pesantren", meta_style),
        Paragraph("Laporan Tagihan", title_style),
        Paragraph(profile.display_address or "Identitas pondok belum dilengkapi.", meta_style),
        Paragraph(" | ".join(part for part in [profile.telepon, profile.wa_admin, profile.email] if part) or " ", meta_style),
        Paragraph(
            f"Digenerate: {timezone.localtime(timezone.now()).strftime('%d %b %Y %H:%M')}",
            meta_style,
        ),
        Spacer(1, 4 * mm),
    ]

    summary_data = [["Ikhtisar", "Nilai"]]
    for label, value in _summary_rows(rows):
        summary_data.append([label, f"{value:,}" if isinstance(value, int) else value])

    summary_table = Table(summary_data, colWidths=[70 * mm, 40 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#A05A2C")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6C8B8")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFF8EF")),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    detail_data = [[
        "Santri",
        "NIS",
        "Kelas",
        "Wali",
        "Jumlah Tagihan",
        "Total Tagihan",
        "Total Dibayar",
        "Sisa",
        "Jatuh Tempo Terakhir",
        "Status",
        "Rincian Tagihan",
    ]]
    for row in rows:
        detail_data.append(
            [
                row["santri"],
                row["nis"],
                row["kelas"],
                row["wali"],
                str(row["bill_count"]),
                f'{row["total_tagihan"]:,}',
                f'{row["total_dibayar"]:,}',
                f'{row["total_sisa"]:,}',
                row["latest_due"].strftime("%d-%m-%Y") if row["latest_due"] else "-",
                row["status"],
                Paragraph(row["detail"].replace("\n", "<br/>"), styles["BodyText"]),
            ]
        )

    detail_table = Table(
        detail_data,
        repeatRows=1,
        colWidths=[30 * mm, 18 * mm, 14 * mm, 28 * mm, 16 * mm, 18 * mm, 18 * mm, 18 * mm, 22 * mm, 16 * mm, 45 * mm],
    )
    detail_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3E2A1F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D6C8B8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (4, 1), (9, -1), "RIGHT"),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFFFFF")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF8EF")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    story.extend([summary_table, Spacer(1, 6 * mm), detail_table])
    doc.build(story)
    return buffer.getvalue()
