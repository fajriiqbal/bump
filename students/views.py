import csv
from datetime import date, datetime
from io import TextIOWrapper
import re

from django.contrib import messages
from django.db import transaction
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView
from accounts.decorators import admin_required
from accounts.utils import log_action

from .forms import StudentForm, StudentImportForm
from .models import Student


class FinanceAdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        user = self.request.user
        return user.is_superuser or getattr(user, "role", "") == "admin"


class StudentListView(LoginRequiredMixin, ListView):
    model = Student
    template_name = "students/list.html"
    context_object_name = "students"


class StudentDetailView(LoginRequiredMixin, DetailView):
    model = Student
    template_name = "students/detail.html"
    context_object_name = "student"


class StudentCreateView(FinanceAdminRequiredMixin, CreateView):
    model = Student
    form_class = StudentForm
    template_name = "students/form.html"
    success_url = reverse_lazy("students:list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["wizard_steps"] = [
            {"title": "Identitas", "hint": "Data dasar santri", "fields": ["nis", "nama_lengkap", "jenis_kelamin", "tempat_lahir", "tanggal_lahir", "foto"]},
            {"title": "Wali", "hint": "Kontak keluarga", "fields": ["nama_ayah", "nama_ibu", "nama_wali", "no_wa_wali"]},
            {"title": "Asrama", "hint": "Penempatan santri", "fields": ["kelas", "kamar", "tanggal_masuk", "tahun_ajaran", "status", "alamat"]},
        ]
        return context


class StudentUpdateView(FinanceAdminRequiredMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = "students/form.html"
    success_url = reverse_lazy("students:list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["wizard_steps"] = [
            {"title": "Identitas", "hint": "Data dasar santri", "fields": ["nis", "nama_lengkap", "jenis_kelamin", "tempat_lahir", "tanggal_lahir", "foto"]},
            {"title": "Wali", "hint": "Kontak keluarga", "fields": ["nama_ayah", "nama_ibu", "nama_wali", "no_wa_wali"]},
            {"title": "Asrama", "hint": "Penempatan santri", "fields": ["kelas", "kamar", "tanggal_masuk", "tahun_ajaran", "status", "alamat"]},
        ]
        return context


class StudentDeleteView(FinanceAdminRequiredMixin, DeleteView):
    model = Student
    template_name = "students/confirm_delete.html"
    success_url = reverse_lazy("students:list")


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value):
    text = _clean_text(value).lower()
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _clean_text(value)
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text).date()
    except ValueError as exc:
        raise ValueError(f"Format tanggal tidak valid: {text}") from exc


def _normalize_gender(value):
    text = _normalize_header(value)
    mapping = {
        "l": Student.Gender.LAKI,
        "lk": Student.Gender.LAKI,
        "laki_laki": Student.Gender.LAKI,
        "laki": Student.Gender.LAKI,
        "pria": Student.Gender.LAKI,
        "p": Student.Gender.PEREMPUAN,
        "pr": Student.Gender.PEREMPUAN,
        "perempuan": Student.Gender.PEREMPUAN,
        "wanita": Student.Gender.PEREMPUAN,
    }
    if text not in mapping:
        raise ValueError("Jenis kelamin harus L/P, Laki-laki, atau Perempuan.")
    return mapping[text]


def _normalize_status(value):
    text = _normalize_header(value)
    mapping = {
        "aktif": Student.Status.AKTIF,
        "nonaktif": Student.Status.NONAKTIF,
        "non_aktif": Student.Status.NONAKTIF,
        "alumni": Student.Status.ALUMNI,
    }
    return mapping.get(text, Student.Status.AKTIF)


def _read_import_rows(uploaded_file):
    filename = (uploaded_file.name or "").lower()
    if filename.endswith(".csv"):
        wrapper = TextIOWrapper(uploaded_file.file, encoding="utf-8-sig")
        reader = csv.DictReader(wrapper)
        normalized_rows = []
        for row in reader:
            normalized = {_normalize_header(key): value for key, value in row.items()}
            if any(_clean_text(value) for value in normalized.values()):
                normalized_rows.append(normalized)
        return normalized_rows

    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_normalize_header(value) for value in rows[0]]
        data_rows = []
        for row in rows[1:]:
            if not any(_clean_text(value) for value in row):
                continue
            data_rows.append(dict(zip(headers, row)))
        return data_rows

    raise ValueError("Format file tidak didukung. Gunakan CSV atau XLSX.")


def _import_students_from_rows(rows, update_existing=True):
    created = 0
    updated = 0
    skipped = 0
    errors = []

    field_map = {
        "nis": "nis",
        "nama_lengkap": "nama_lengkap",
        "jenis_kelamin": "jenis_kelamin",
        "tempat_lahir": "tempat_lahir",
        "tanggal_lahir": "tanggal_lahir",
        "alamat": "alamat",
        "nama_ayah": "nama_ayah",
        "nama_ibu": "nama_ibu",
        "nama_wali": "nama_wali",
        "no_wa_wali": "no_wa_wali",
        "kelas": "kelas",
        "kamar": "kamar",
        "tanggal_masuk": "tanggal_masuk",
        "tahun_ajaran": "tahun_ajaran",
        "status": "status",
    }

    with transaction.atomic():
        for index, row in enumerate(rows, start=2):
            try:
                normalized = {}
                for source, target in field_map.items():
                    normalized[target] = _clean_text(row.get(source))

                nis = normalized["nis"]
                nama_lengkap = normalized["nama_lengkap"]
                if not nis or not nama_lengkap:
                    raise ValueError("NIS dan nama lengkap wajib diisi.")

                payload = {
                    "nama_lengkap": nama_lengkap,
                    "jenis_kelamin": _normalize_gender(row.get("jenis_kelamin")),
                    "tempat_lahir": normalized["tempat_lahir"],
                    "tanggal_lahir": _parse_date(row.get("tanggal_lahir")),
                    "alamat": normalized["alamat"],
                    "nama_ayah": normalized["nama_ayah"],
                    "nama_ibu": normalized["nama_ibu"],
                    "nama_wali": normalized["nama_wali"],
                    "no_wa_wali": normalized["no_wa_wali"],
                    "kelas": normalized["kelas"],
                    "kamar": normalized["kamar"],
                    "tanggal_masuk": _parse_date(row.get("tanggal_masuk")),
                    "tahun_ajaran": normalized["tahun_ajaran"],
                    "status": _normalize_status(row.get("status")),
                }

                if update_existing:
                    _, was_created = Student.objects.update_or_create(nis=nis, defaults=payload)
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    if Student.objects.filter(nis=nis).exists():
                        skipped += 1
                        continue
                    Student.objects.create(nis=nis, **payload)
                    created += 1
            except Exception as exc:
                skipped += 1
                errors.append(f"Baris {index}: {exc}")

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


@admin_required
def export_students_xlsx(request):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Santri"
    headers = ["NIS", "Nama Lengkap", "Jenis Kelamin", "Kelas", "Kamar", "Status"]
    ws.append(headers)
    for s in Student.objects.all():
        ws.append([s.nis, s.nama_lengkap, s.get_jenis_kelamin_display(), s.kelas, s.kamar, s.status])
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="data_santri.xlsx"'
    wb.save(response)
    log_action(request.user, "export_students_xlsx", "Student")
    return response


@admin_required
def import_students(request):
    form = StudentImportForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        try:
            rows = _read_import_rows(form.cleaned_data["file"])
            result = _import_students_from_rows(
                rows,
                update_existing=form.cleaned_data["update_existing"],
            )
        except ValueError as exc:
            form.add_error("file", str(exc))
        else:
            messages.success(
                request,
                f"Import selesai: {result['created']} data baru, {result['updated']} diperbarui, {result['skipped']} dilewati.",
            )
            if result["errors"]:
                preview = "; ".join(result["errors"][:5])
                if len(result["errors"]) > 5:
                    preview += " ..."
                messages.warning(request, f"Ada beberapa baris yang dilewati: {preview}")
            log_action(request.user, "import_students", "Student")
            return redirect("students:list")

    sample_headers = [
        "nis",
        "nama_lengkap",
        "jenis_kelamin",
        "tempat_lahir",
        "tanggal_lahir",
        "kelas",
        "kamar",
        "status",
    ]
    sample_rows = [
        ["2024001", "Ahmad Fauzi", "L", "Surabaya", "2008-01-12", "7A", "01", "aktif"],
        ["2024002", "Siti Aisyah", "P", "Sidoarjo", "2008-05-20", "7A", "02", "aktif"],
    ]
    return render(
        request,
        "students/import.html",
        {
            "form": form,
            "sample_headers": sample_headers,
            "sample_rows": sample_rows,
        },
    )


@admin_required
def import_students_csv(request):
    return import_students(request)
